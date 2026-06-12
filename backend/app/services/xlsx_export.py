"""Exportação de simulações em XLSX (openpyxl) e CSV."""

import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import Simulation, SimulationItem

ITEM_HEADERS = [
    "Item", "Tipo", "Ano", "Receita anual", "Carga atual", "Créditos atuais", "Carga atual líquida",
    "CBS", "IBS", "IS", "Tributos legados", "Créditos futuros", "Carga futura líquida",
    "Δ carga", "Margem atual %", "Margem futura %", "Preço de equilíbrio", "Impacto caixa (split)",
]


def _item_row(item: SimulationItem) -> list:
    return [
        item.item_name, "Produto" if item.item_kind == "product" else "Serviço", item.year,
        float(item.annual_revenue), float(item.current_tax_total), float(item.current_credits),
        float(item.current_net_burden), float(item.future_cbs), float(item.future_ibs),
        float(item.future_is), float(item.future_legacy), float(item.future_credits),
        float(item.future_net_burden), float(item.delta_net),
        float(item.margin_current_pct) if item.margin_current_pct is not None else None,
        float(item.margin_future_pct) if item.margin_future_pct is not None else None,
        float(item.breakeven_price) if item.breakeven_price is not None else None,
        float(item.cash_flow_impact),
    ]


def build_simulation_xlsx(simulation: Simulation, items: list[SimulationItem]) -> bytes:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="102A43")
    header_font = Font(color="FFFFFF", bold=True)

    ws = wb.active
    ws.title = "Resumo por ano"
    summary_headers = ["Ano", "Receita", "Carga atual", "Carga futura", "Δ", "Δ % receita",
                       "CBS", "IBS", "IS", "Legado", "Créditos atuais", "Créditos futuros", "Caixa (split)"]
    ws.append(summary_headers)
    for year in (simulation.summary or {}).get("years", []):
        ws.append([
            year["year"], year["revenue"], year["current_total"], year["future_total"], year["delta"],
            year["delta_pct_revenue"], year["cbs"], year["ibs"], year["is"], year["legacy"],
            year["credits_current"], year["credits_future"], year["cash_flow_impact"],
        ])

    ws_items = wb.create_sheet("Itens")
    ws_items.append(ITEM_HEADERS)
    for item in items:
        ws_items.append(_item_row(item))

    ws_premises = wb.create_sheet("Premissas")
    ws_premises.append(["Premissa", "Valor"])
    ws_premises.append(["Simulação", simulation.name])
    ws_premises.append(["Cenário", simulation.scenario])
    for key, value in (simulation.assumptions_snapshot or {}).items():
        ws_premises.append([key, str(value)])
    ws_premises.append([
        "Aviso",
        "Estimativas baseadas em premissas configuráveis; não substituem parecer profissional.",
    ])

    for sheet in (ws, ws_items, ws_premises):
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for column_index in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(column_index)].width = 16
        sheet.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_simulation_csv(items: list[SimulationItem]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(ITEM_HEADERS)
    for item in items:
        writer.writerow(_item_row(item))
    return output.getvalue().encode("utf-8-sig")
